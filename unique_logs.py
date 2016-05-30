import xlrd  					#Excel module
import re 						#Parser module
import os
import glob
import xlwt
import openpyxl
import codecs
dir_list =[]
for filename in os.listdir("C:/work/ktk/ACL/Script/Events_WIN/"):	
	dir_list.insert (0,filename)
files_counter = len (dir_list)
files_regex_system = ".*system.*"
files_regex_application = ".*application.*"
file_regex = "^M42.*|^P25.*|^M0042.*|^P0025.*"
counter = 0
catalog =''
clean_file_name_list =[]
inside_file_list = []
row_id_list = []
while files_counter > 0:
	find_file_name = re.findall (file_regex,dir_list[counter])	
	if find_file_name != []:
		for line in open('C:/work/ktk/ACL/Script/Events_WIN/{0}'.format(find_file_name[0]), encoding='utf-8' ).readlines():
			inside_file_list.insert(0,line.strip())
		len_inside_file_list = len (inside_file_list)		
		if len_inside_file_list > 3:			
			clean_file_name_list.insert (0,find_file_name[0])
		inside_file_list = []
		print (clean_file_name_list)
	counter += 1
	files_counter -= 1
#print ('VISHLI')
new_wb = openpyxl.load_workbook(filename = 'C:/work/ktk/ACL/Script/Events_WIN/example.xlsx')
sheet = new_wb['ALL Events']
#print (clean_file_name_list)
length_clean_file_name_list = len (clean_file_name_list)
counter_for_files = 0
while length_clean_file_name_list > 0:
	new_list = []
	for line in open('C:/work/ktk/ACL/Script/Events_WIN/{0}'.format(clean_file_name_list[counter_for_files]), encoding='utf-8' ).readlines():
		new_list.append(line.strip())
	find_system = re.findall (files_regex_system,clean_file_name_list[counter_for_files])
	find_application = re.findall (files_regex_application,clean_file_name_list[counter_for_files])
	if find_system != []:
		catalog = 'System'
	if find_application != []:
		catalog = 'Application'		
	test_regex_mes = '^Message            :'
	test_regex_source = '^Source             :'
	regex_event_id = "^EventID .*$"
	regex_machine_name = "^MachineName .*$"
	regex_exact_machine_name = "M42-.*$|P25-.*$|M0042-.*$|P0025-.*$"
	regex_entry_type = "^EntryType.*$"
	regex_entry_type_value = "Error$|Critical$"
	regex_source = "^Source             :.*$"
	regex_delete_source = "Source             : "
	regex_time_stamp = "^TimeGenerated      : .*"
	machine_list =[]
	event_list =[]
	entry_type_list =[]
	source_list = []
	message_list =[]
	source_msg = []
	time_list = []
	x = 0
	y = len(new_list)
	length_entry_type = 0
	length_machine = 0
	length_event = 0
	length_source = 0
	length_message = 0
	length_source_msg = 0
	length_time = 0	
	while x < y:
		find_event = re.findall(regex_event_id, new_list[x])
		find_time = re.findall(regex_time_stamp, new_list[x])
		find_machine_name = re.findall(regex_machine_name, new_list[x])
		find_entry_type = re.findall (regex_entry_type, new_list[x])
		find_source = re.findall (regex_source, new_list[x])
		find_message = re.findall (test_regex_mes,new_list[x])
		find_source_msg = re.findall (test_regex_source,new_list[x])
		if find_source_msg != []:
			source_msg.insert(length_source_msg,x)
			length_source_msg += 1
		if find_message != []:
			message_list.insert(length_message,x)
			length_message += 1
		if find_source != []:
			source_list.insert(length_source,find_source[0])
			length_source += 1
		if find_entry_type != []:		
			entry_type_list.insert(length_entry_type,find_entry_type[0])
			length_entry_type += 1
		if find_machine_name != []:
			machine_list.insert (length_machine,find_machine_name[0])
			length_machine += 1
		if find_event != []:
			event_list.insert (length_event,find_event[0])
			length_event += 1
		if find_time != []:
			time_list.insert (length_time,find_time[0])
			length_time += 1
		x+=1
	event_list_len = len(event_list)
	row_id_list.insert (0,event_list_len)
	#print (event_list_len)
	#print (row_id_list)
	print (time_list)
	row_id_length = len (row_id_list)
	entry_type_list_len = len(entry_type_list)
	
	entry_counter = 0		
	last_entry_type_list = []
	# while entry_counter < entry_type_list_len:
	# 	exact_entry_type = re.findall (regex_entry_type_value,entry_type_list[entry_counter])	
	# 	last_entry_type_list.insert (entry_counter,exact_entry_type[0])
	# 	mes_value = message_list[entry_counter]
	# 	source_value = source_msg[entry_counter]
	# 	sheet.cell(row=new_row_counter, column=6).value = new_list[mes_value]
	# 	new_source_value = source_value
	# 	new_mes_value = mes_value
	# 	if source_value - mes_value > 1:
	# 		while mes_value < source_value:				
	# 			mes_value += 1
	# 			new = sheet.cell(row=new_row_counter, column=6).value + new_list[mes_value]
	# 			sheet.cell(row=new_row_counter, column=6).value=new	
	# 	new_row_counter += 1
	# 	entry_counter += 1
	#print (machine_list)	
	exact_machine_name = re.findall(regex_exact_machine_name, machine_list[0])		
	if row_id_length == 1:
		row_counter = 1
		entry_type_count = 0
		optional_counter = 0
		z = 0
		new_row_counter = 1	
		while z < event_list_len:
			while entry_counter < entry_type_list_len:
				exact_entry_type = re.findall (regex_entry_type_value,entry_type_list[entry_counter])	
				last_entry_type_list.insert (entry_counter,exact_entry_type[0])
				mes_value = message_list[entry_counter]
				source_value = source_msg[entry_counter]
				sheet.cell(row=new_row_counter, column=8).value = new_list[mes_value]
				new_source_value = source_value
				new_mes_value = mes_value
				if source_value - mes_value > 1:
					while mes_value < source_value:				
						mes_value += 1
						new = sheet.cell(row=new_row_counter, column=8).value + new_list[mes_value]
						sheet.cell(row=new_row_counter, column=8).value=new	
				new_row_counter += 1
				entry_counter += 1
			clear_source = re.sub (regex_delete_source,'',source_list[z])
			sheet.cell(row=row_counter, column=6).value = event_list[z]
			sheet.cell(row=row_counter, column=1).value = exact_machine_name[0]
			sheet.cell(row=row_counter, column=4).value = catalog
			sheet.cell(row=row_counter, column=2).value = time_list[z]
			sheet.cell(row=row_counter, column=3).value = last_entry_type_list[z]
			sheet.cell(row=row_counter, column=5).value = clear_source
			z+=1
			row_counter += 1
			#print (row_counter)
			optional_counter += 1			
	if row_id_length > 1:
		v = 0
		#print (row_counter)
		while v < event_list_len:
			while entry_counter < entry_type_list_len:
				exact_entry_type = re.findall (regex_entry_type_value,entry_type_list[entry_counter])	
				last_entry_type_list.insert (entry_counter,exact_entry_type[0])
				mes_value = message_list[entry_counter]
				source_value = source_msg[entry_counter]
				sheet.cell(row=new_row_counter, column=8).value = new_list[mes_value]
				new_source_value = source_value
				new_mes_value = mes_value
				if source_value - mes_value > 1:
					while mes_value < source_value:				
						mes_value += 1
						new = sheet.cell(row=new_row_counter, column=8).value + new_list[mes_value]
						sheet.cell(row=new_row_counter, column=8).value=new	
				new_row_counter += 1
				entry_counter += 1
			clear_source = re.sub (regex_delete_source,'',source_list[v])
			sheet.cell(row=row_counter, column=6).value = event_list[v]
			sheet.cell(row=row_counter, column=1).value = exact_machine_name[0]
			sheet.cell(row=row_counter, column=4).value = catalog
			sheet.cell(row=row_counter, column=2).value = time_list[v]
			sheet.cell(row=row_counter, column=3).value = last_entry_type_list[v]
			sheet.cell(row=row_counter, column=5).value = clear_source			
			row_counter += 1
			v += 1
			#print (row_counter)
	length_clean_file_name_list -= 1
	counter_for_files += 1
	print ('END OF CYCLE')
	#print (event_list)
	#print (row_counter)
new_wb.save('C:/work/ktk/ACL/Script/Events_WIN/example.xlsx')
